import json
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from tender_review.approval_optimizer import (
    ApprovalOptimizer,
    TaskArtifactResolver,
    group_approval_cases,
    load_approval_dataset,
    summarize_group_checks,
)
from tender_review.config import Settings


def write_json(path: Path, value) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(value, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )


def approval_payload(file_name, status, task_id, product_id, opinions):
    return {
        "source": {
            "file_name": file_name,
            "relative_path": f"category/{file_name}",
        },
        "review": {
            "status": status,
            "task_id": task_id,
            "product_id": product_id,
            "opinions": opinions,
            "missing_review_items": ["7"] if status == "partial" else [],
        },
    }


def opinion(item, point_id, compliant, text):
    return {
        "review_item": item,
        "review_point_id": point_id,
        "compliant": compliant,
        "opinion": text,
        "evidence": {"page_number": 1},
        "reasoning": "人工判断依据",
        "review_point": "原始要点",
    }


class FakePlatform:
    def __init__(self):
        self.refreshed = 0
        self.calls = []

    def refresh_token(self):
        self.refreshed += 1

    def workflow_test(self, payload):
        self.calls.append(payload)
        return {"point": payload["point"], "task_id": payload["taskId"]}


class FakeLlm:
    def __init__(self):
        self.optimize_calls = 0

    def evaluate_approval_case(
        self, *, expected_compliant, expected_opinions, platform_output
    ):
        point = platform_output["point"]
        if expected_compliant:
            passed = point != "候选一"
        else:
            passed = point in {"候选一", "候选二"}
        return {
            "passed": passed,
            "observed_compliant": expected_compliant and passed,
            "covered_opinions": expected_opinions if passed else [],
            "missed_opinions": [] if passed else expected_opinions,
            "reason": "通过" if passed else "未通过",
        }

    def evaluate_approval_group(self, cases):
        return {
            case["case_id"]: {
                "case_id": case["case_id"],
                **self.evaluate_approval_case(
                    expected_compliant=case["expected_compliant"],
                    expected_opinions=case["expected_opinions"],
                    platform_output=case["platform_output"],
                ),
            }
            for case in cases
        }

    def optimize_review_point_group(self, **kwargs):
        self.optimize_calls += 1
        candidate = "候选一" if self.optimize_calls == 1 else "候选二"
        return {
            "optimized_point": candidate,
            "change_note": f"第{self.optimize_calls}轮",
            "coverage_strategy": "联合覆盖",
        }


class ApprovalInputTests(unittest.TestCase):
    def test_real_evaluation_payload_drops_large_extracted_content(self):
        from tender_review.approval_optimizer import _compact_platform_output

        compact = _compact_platform_output(
            {
                "ai_result": {"result": "false"},
                "reasoning": "明确问题",
                "case_result": {"result": "false"},
                "content_result": "x" * 200000,
            }
        )

        self.assertEqual(
            set(compact), {"ai_result", "reasoning", "case_result"}
        )
        self.assertNotIn("content_result", compact)

    def test_loads_partial_opinions_and_groups_by_review_item(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            source = root / "pdf"
            opinions_dir = root / "opinions"
            (source / "category").mkdir(parents=True)
            for name in ("a.pdf", "b.pdf", "c.pdf"):
                (source / "category" / name).write_bytes(b"pdf")

            write_json(
                opinions_dir / "a.json",
                approval_payload(
                    "a.pdf",
                    "completed",
                    "task-a",
                    "product",
                    [
                        opinion("41", "point-a-41", False, "问题一"),
                        opinion("41", "point-a-41", False, "问题二"),
                        opinion("7", "point-a-7", True, "符合要求"),
                    ],
                ),
            )
            write_json(
                opinions_dir / "b.json",
                approval_payload(
                    "b.pdf",
                    "partial",
                    "task-b",
                    "product",
                    [opinion("41", "point-b-41", True, "符合要求")],
                ),
            )
            write_json(
                opinions_dir / "c.json",
                approval_payload("c.pdf", "missing", "pending", "", []),
            )

            dataset = load_approval_dataset(opinions_dir, source)
            groups = group_approval_cases(dataset.cases)

            self.assertEqual(len(dataset.cases), 3)
            self.assertEqual(set(groups), {"7", "41"})
            self.assertEqual(len(groups["41"]), 2)
            target = next(case for case in groups["41"] if not case.expected_compliant)
            self.assertEqual(target.expected_opinions, ("问题一", "问题二"))
            self.assertEqual(len(dataset.skipped_files), 1)

    def test_group_acceptance_requires_targets_and_protection_in_same_run(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            source = root / "pdf"
            opinions_dir = root / "opinions"
            source.mkdir()
            (source / "a.pdf").write_bytes(b"pdf")
            (source / "b.pdf").write_bytes(b"pdf")
            write_json(
                opinions_dir / "a.json",
                approval_payload(
                    "a.pdf",
                    "completed",
                    "task-a",
                    "product",
                    [opinion("41", "point-a", False, "必须检出")],
                ),
            )
            write_json(
                opinions_dir / "b.json",
                approval_payload(
                    "b.pdf",
                    "completed",
                    "task-b",
                    "product",
                    [opinion("41", "point-b", True, "保持合规")],
                ),
            )
            cases = load_approval_dataset(opinions_dir, source).cases
            results = [
                {
                    "case_id": cases[0].case_id,
                    "checks": [{"evaluation": {"passed": True}}],
                },
                {
                    "case_id": cases[1].case_id,
                    "checks": [{"evaluation": {"passed": False}}],
                },
            ]

            summary = summarize_group_checks(cases, results, stability_runs=1)

            self.assertFalse(summary["all_passed"])
            self.assertTrue(summary["target_cases_passed"])
            self.assertFalse(summary["protection_cases_passed"])


class ApprovalOptimizerFlowTests(unittest.TestCase):
    def test_rejects_regression_then_writes_shared_accepted_candidate(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            source = root / "pdf"
            opinions_dir = root / "opinions"
            batch_run = root / "runs" / "batch_test"
            run_dir = root / "runs" / "approval_test"
            source.mkdir(parents=True)
            (source / "a.pdf").write_bytes(b"pdf")
            (source / "b.pdf").write_bytes(b"pdf")

            excel = root / "rules.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["工作流ID", "工作流名称", "审核文件名", "审评项目", "审评要点"])
            sheet.append(["workflow", "招标审核", "招标文件", "41", "原始要点"])
            workbook.save(excel)

            write_json(
                opinions_dir / "a.json",
                approval_payload(
                    "a.pdf",
                    "completed",
                    "task-a",
                    "product",
                    [opinion("41", "point-a", False, "必须检出问题")],
                ),
            )
            write_json(
                opinions_dir / "b.json",
                approval_payload(
                    "b.pdf",
                    "completed",
                    "task-b",
                    "product",
                    [opinion("41", "point-b", True, "保持合规")],
                ),
            )

            records = []
            for index, (name, task_id, point_id) in enumerate(
                (("a.pdf", "task-a", "point-a"), ("b.pdf", "task-b", "point-b")),
                start=1,
            ):
                task_dir = batch_run / "tasks" / f"{index:03d}"
                write_json(
                    task_dir / "01_上传ZIP_响应.json",
                    {
                        "reviewPoints": [
                            {
                                "id": point_id,
                                "reviewItem": "41",
                                "workFlowId": "workflow",
                                "file": "招标文件",
                                "content": "{}",
                                "point": "原始要点",
                            }
                        ]
                    },
                )
                records.append(
                    {
                        "file_name": name,
                        "task_id": task_id,
                        "successful_task_id": task_id,
                        "task_dir": str(task_dir),
                    }
                )
            write_json(batch_run / "03_任务索引.json", records)

            settings = Settings(
                base_url="https://example.invalid",
                username="user",
                llm_url="https://example.invalid/llm",
                llm_api_key="test",
                llm_model="fake",
                excel_path=excel,
                pdf_path=source / "a.pdf",
                runs_dir=root / "runs",
            )
            platform = FakePlatform()
            llm = FakeLlm()
            optimizer = ApprovalOptimizer(
                settings=settings,
                opinion_dir=opinions_dir,
                source_dir=source,
                batch_run=batch_run,
                max_iterations=2,
                stability_runs=1,
                run_dir=run_dir,
                platform=platform,
                llm=llm,
            )

            summary = optimizer.run()

            self.assertEqual(summary["group_status_counts"], {"optimized": 1})
            self.assertEqual(summary["optimized_review_items"], ["41"])
            self.assertEqual(llm.optimize_calls, 2)
            first_attempt = json.loads(
                (run_dir / "items" / "41" / "02_第01轮.json").read_text(
                    encoding="utf-8"
                )
            )
            self.assertFalse(first_attempt["accepted"])
            self.assertFalse(
                first_attempt["evaluation"]["summary"]["protection_cases_passed"]
            )
            updated = load_workbook(
                run_dir / "审核要点_优化结果_最终.xlsx", read_only=True, data_only=True
            )
            try:
                self.assertEqual(updated.active.cell(2, 5).value, "候选二")
            finally:
                updated.close()
            self.assertTrue((run_dir / "审核要点_优化结果_v001.xlsx").is_file())
            self.assertEqual(platform.refreshed, 1)

    def test_resolver_uses_exact_review_point_metadata(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            source = root / "pdf"
            opinions_dir = root / "opinions"
            batch_run = root / "batch"
            task_dir = batch_run / "tasks" / "001"
            source.mkdir()
            (source / "a.pdf").write_bytes(b"pdf")
            write_json(
                opinions_dir / "a.json",
                approval_payload(
                    "a.pdf",
                    "completed",
                    "task-a",
                    "product",
                    [opinion("41", "wanted", False, "问题")],
                ),
            )
            write_json(
                task_dir / "01_上传ZIP_响应.json",
                {
                    "reviewPoints": [
                        {
                            "id": "wanted",
                            "reviewItem": "41",
                            "workFlowId": "workflow",
                            "file": "招标文件",
                            "content": "{}",
                            "point": "原始",
                        }
                    ]
                },
            )
            write_json(
                batch_run / "03_任务索引.json",
                [
                    {
                        "file_name": "a.pdf",
                        "task_id": "task-a",
                        "task_dir": str(task_dir),
                    }
                ],
            )
            case = load_approval_dataset(opinions_dir, source).cases[0]

            context = TaskArtifactResolver(batch_run).resolve(case)

            self.assertEqual(context.review_point_id, "wanted")
            self.assertEqual(context.workflow_id, "workflow")


if __name__ == "__main__":
    unittest.main()
