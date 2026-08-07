import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from tender_review.rules import (
    load_review_rule,
    load_review_rules,
    write_optimized_excel,
    write_optimized_excel_updates,
)


class RuleWorkbookTests(unittest.TestCase):
    def test_load_and_update_one_rule(self):
        with tempfile.TemporaryDirectory() as directory:
            source = Path(directory) / "rules.xlsx"
            target = Path(directory) / "optimized.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["工作流ID", "工作流名称", "审核文件名", "审评项目", "审评要点"])
            sheet.append(["wf", "workflow", "招标文件", "41", "原始要点"])
            workbook.save(source)

            rule = load_review_rule(source, "41")
            self.assertEqual(rule.point, "原始要点")
            write_optimized_excel(source, target, "41", "优化要点")

            updated = load_workbook(target, read_only=True, data_only=True)
            try:
                self.assertEqual(updated.active.cell(2, 5).value, "优化要点")
            finally:
                updated.close()

    def test_load_and_update_multiple_rules(self):
        with tempfile.TemporaryDirectory() as directory:
            source = Path(directory) / "rules.xlsx"
            target = Path(directory) / "optimized.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["工作流ID", "工作流名称", "审核文件名", "审评项目", "审评要点"])
            sheet.append(["wf", "workflow", "招标文件", "7", "要点7"])
            sheet.append(["wf", "workflow", "招标文件", "41", "要点41"])
            workbook.save(source)

            rules = load_review_rules(source)
            write_optimized_excel_updates(
                source, target, {"7": "优化7", "41": "优化41"}
            )

            self.assertEqual(set(rules), {"7", "41"})
            updated = load_workbook(target, read_only=True, data_only=True)
            try:
                self.assertEqual(updated.active.cell(2, 5).value, "优化7")
                self.assertEqual(updated.active.cell(3, 5).value, "优化41")
            finally:
                updated.close()


if __name__ == "__main__":
    unittest.main()
