from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

from tender_review.rule_management.public import (
    CompleteEvaluationGate,
    CreateRuleVersion,
    EvaluationGate,
    EvaluationGateStatus,
    InMemoryRuleVersionRepository,
    PublishRuleVersion,
    RollbackRuleSet,
    RuleDiffChange,
    RuleProvenance,
    RuleSet,
    RuleVersion,
    RuleVersionDiff,
    RuleVersionRepository,
    RuleVersionService,
    RuleVersionStatus,
    canonical_json,
)


__all__ = [
    "CompleteEvaluationGate",
    "CreateRuleVersion",
    "EvaluationGate",
    "EvaluationGateStatus",
    "InMemoryRuleVersionRepository",
    "PublishRuleVersion",
    "ReviewRule",
    "RollbackRuleSet",
    "RuleDiffChange",
    "RuleProvenance",
    "RuleSet",
    "RuleVersion",
    "RuleVersionDiff",
    "RuleVersionRepository",
    "RuleVersionService",
    "RuleVersionStatus",
    "canonical_json",
    "load_review_rule",
    "load_review_rules",
    "write_optimized_excel",
    "write_optimized_excel_updates",
]


@dataclass(frozen=True)
class ReviewRule:
    row_number: int
    workflow_id: str
    workflow_name: str
    file_name: str
    review_item: str
    point: str


def load_review_rule(excel_path: Path, review_item: str) -> ReviewRule:
    workbook = load_workbook(excel_path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        headers = {
            str(cell.value).strip(): index
            for index, cell in enumerate(sheet[1], start=1)
            if cell.value is not None
        }
        required = ["工作流ID", "工作流名称", "审核文件名", "审评项目", "审评要点"]
        missing = [name for name in required if name not in headers]
        if missing:
            raise ValueError(f"Excel is missing columns: {', '.join(missing)}")

        for row_number in range(2, sheet.max_row + 1):
            value = sheet.cell(row_number, headers["审评项目"]).value
            if str(value).strip() != str(review_item).strip():
                continue
            return ReviewRule(
                row_number=row_number,
                workflow_id=str(sheet.cell(row_number, headers["工作流ID"]).value or "").strip(),
                workflow_name=str(sheet.cell(row_number, headers["工作流名称"]).value or "").strip(),
                file_name=str(sheet.cell(row_number, headers["审核文件名"]).value or "").strip(),
                review_item=str(value).strip(),
                point=str(sheet.cell(row_number, headers["审评要点"]).value or "").strip(),
            )
    finally:
        workbook.close()
    raise ValueError(f"Review item {review_item!r} was not found in {excel_path}")


def load_review_rules(excel_path: Path) -> dict[str, ReviewRule]:
    workbook = load_workbook(excel_path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        headers = {
            str(cell.value).strip(): index
            for index, cell in enumerate(sheet[1], start=1)
            if cell.value is not None
        }
        required = ["工作流ID", "工作流名称", "审核文件名", "审评项目", "审评要点"]
        missing = [name for name in required if name not in headers]
        if missing:
            raise ValueError(f"Excel is missing columns: {', '.join(missing)}")

        rules: dict[str, ReviewRule] = {}
        for row_number in range(2, sheet.max_row + 1):
            value = str(
                sheet.cell(row_number, headers["审评项目"]).value or ""
            ).strip()
            if not value:
                continue
            if value in rules:
                raise ValueError(f"Duplicate review item {value!r} in {excel_path}")
            rules[value] = ReviewRule(
                row_number=row_number,
                workflow_id=str(
                    sheet.cell(row_number, headers["工作流ID"]).value or ""
                ).strip(),
                workflow_name=str(
                    sheet.cell(row_number, headers["工作流名称"]).value or ""
                ).strip(),
                file_name=str(
                    sheet.cell(row_number, headers["审核文件名"]).value or ""
                ).strip(),
                review_item=value,
                point=str(
                    sheet.cell(row_number, headers["审评要点"]).value or ""
                ).strip(),
            )
        return rules
    finally:
        workbook.close()


def write_optimized_excel(
    source: Path, target: Path, review_item: str, optimized_point: str
) -> None:
    workbook = load_workbook(source)
    sheet = workbook.active
    headers = {
        str(cell.value).strip(): index
        for index, cell in enumerate(sheet[1], start=1)
        if cell.value is not None
    }
    for row_number in range(2, sheet.max_row + 1):
        value = sheet.cell(row_number, headers["审评项目"]).value
        if str(value).strip() == str(review_item).strip():
            sheet.cell(row_number, headers["审评要点"]).value = optimized_point
            workbook.save(target)
            workbook.close()
            return
    workbook.close()
    raise ValueError(f"Review item {review_item!r} was not found in {source}")


def write_optimized_excel_updates(
    source: Path, target: Path, optimized_points: dict[str, str]
) -> None:
    workbook = load_workbook(source)
    try:
        sheet = workbook.active
        headers = {
            str(cell.value).strip(): index
            for index, cell in enumerate(sheet[1], start=1)
            if cell.value is not None
        }
        required = ["审评项目", "审评要点"]
        missing = [name for name in required if name not in headers]
        if missing:
            raise ValueError(f"Excel is missing columns: {', '.join(missing)}")

        remaining = {str(key).strip() for key in optimized_points}
        for row_number in range(2, sheet.max_row + 1):
            review_item = str(
                sheet.cell(row_number, headers["审评项目"]).value or ""
            ).strip()
            if review_item not in optimized_points:
                continue
            point = str(optimized_points[review_item] or "").strip()
            if not point:
                raise ValueError(
                    f"Optimized review point {review_item!r} must not be empty"
                )
            sheet.cell(row_number, headers["审评要点"]).value = point
            remaining.discard(review_item)

        if remaining:
            raise ValueError(
                "Review items were not found in Excel: " + ", ".join(sorted(remaining))
            )
        target.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(target)
    finally:
        workbook.close()
